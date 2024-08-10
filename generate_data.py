import os
import random

import django
from openpyxl import load_workbook
from faker import Faker
from django.db import transaction

# Инициализация Faker
fake = Faker()

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'calculation_cost.settings')
django.setup()
from calculation.models import TaxRate, MarketplaceCommission, Product


# Чтение файла с комиссиями маркетплейса из Excel
def load_marketplace_commissions(_filepath, _num_products):
    wb = load_workbook(filename=_filepath)
    sheet = wb.active
    commissions = []

    # Пропускаем заголовок и читаем данные
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if _num_products == 0:
            break
        commissions.append({
            'category': row[0],  # Категория
            'item': row[1],  # Предмет
            'fbo_rate': float(str(row[2]).replace(',', '.')),  # FBO
            'fbs_rate': float(str(row[3]).replace(',', '.')),  # FBS
            'dbs_rate': float(str(row[4]).replace(',', '.')),  # DBS
        })
        _num_products -= 1

    return commissions


# Генерация данных для моделей
@transaction.atomic
def generate_data(_filepath, num_products):
    commissions = load_marketplace_commissions(_filepath, num_products)

    tax_rates = [
        TaxRate.objects.get_or_create(rate=rate)[0]
        for rate, _ in TaxRate._meta.get_field('rate').choices
    ]

    for commission in commissions:
        marketplace_commission, _ = MarketplaceCommission.objects.get_or_create(
            category=commission['category'],
            fbs_rate=commission['fbs_rate'],
            fbo_rate=commission['fbo_rate'],
            dbs_rate=commission['dbs_rate'],
        )

        desired_od_margin = random.choice([True, False])
        desired_price = None
        margin_percentage = None

        purchase_price = round(random.uniform(100, 10000), 2)
        if desired_od_margin:
            desired_price = round(random.uniform(purchase_price, 20000), 2)
        else:
            margin_percentage = round(random.uniform(1, 20), 2)

        product = Product.objects.create(
            name=commission['item'],
            purchase_price=purchase_price,
            desired_price=desired_price,
            margin_percentage=margin_percentage,
            logistics_cost=1000.0,
            logistics_quantity=100,
            tax_rate=random.choice(tax_rates),
            category=marketplace_commission,
            height=random.choice([10.0, 3.0, 1.0, 20.0]),
            length=random.choice([10.0, 3.0, 1.0, 20.0]),
            width=random.choice([5.0, 3.0, 1.0, 10.0]),
        )
        print(f"Created product: {product.name}")


if __name__ == "__main__":
    filepath = 'сomission.xlsx'
    generate_data(filepath, num_products=2000)
